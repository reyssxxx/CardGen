"""
Сервис импорта оценок из Excel и генерации шаблонов.
"""
import difflib
import io
import logging
from datetime import datetime
from typing import Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logger = logging.getLogger(__name__)


VALID_GRADES = {'1', '2', '3', '4', '5', 'н', 'б'}


def parse_grades_excel(file_path: str, class_name: str, valid_students: list) -> dict:
    """
    Парсит Excel-файл с оценками.

    Формат файла:
      Строка 1: "Период:" | "ДД.ММ.ГГГГ" | "ДД.ММ.ГГГГ"  (начало и конец)
      Строка 2 (заголовки): Предмет | Ученик1 | Ученик2 | ...
      Строки 3+:            Математика | 5 4 | 3 | ...

    Несколько оценок в ячейке — разделяются пробелами.
    Пустая ячейка = нет оценок.
    Все оценки привязываются к дате начала периода.

    Returns:
        {
            'grades': [{'student_name', 'class', 'subject', 'grade', 'date'}, ...],
            'skipped': ['Имя которое не найдено', ...],
            'period_start': datetime,
            'period_end': datetime,
            'count': int,
        }
    """
    valid_lower = {s.lower(): s for s in valid_students}

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise ValueError("Файл пустой или слишком короткий")

    # ── Строка 1: период ────────────────────────────────────────────────────
    period_row = rows[0]
    period_start = period_end = None
    for cell in period_row[1:]:
        if cell is None:
            continue
        parsed = _parse_date(cell)
        if parsed and period_start is None:
            period_start = parsed
        elif parsed and period_end is None:
            period_end = parsed
            break

    if not period_start:
        raise ValueError(
            "Не найден период в строке 1. "
            "Ожидается: ячейки B1 и C1 с датами ДД.ММ.ГГГГ"
        )
    if not period_end:
        period_end = period_start

    grade_date_str = period_start.strftime('%d.%m.%Y')

    # ── Строка 2: заголовки с именами учеников ──────────────────────────────
    header = rows[1]
    if len(header) < 2:
        raise ValueError("Неверный формат строки 2: нужно минимум 2 столбца (Предмет, Ученик...)")

    student_names_raw = [str(c).strip() for c in header[1:] if c is not None]
    if not student_names_raw:
        raise ValueError("Не найдено ни одного ученика в строке 2")

    matched_students = []
    skipped = []
    for i, raw in enumerate(student_names_raw):
        matched = _match_name(raw, valid_lower)
        if matched:
            matched_students.append((i, matched))
        else:
            if raw not in skipped:
                skipped.append(raw)
                logger.debug("Name not matched in header: %r", raw)

    # ── Строки 3+: предметы и оценки ────────────────────────────────────────
    grades_list = []
    for row in rows[2:]:
        if not row or row[0] is None:
            continue
        subject = str(row[0]).strip()
        if not subject:
            continue

        for col_i, student_name in matched_students:
            cell_idx = col_i + 1
            cell_value = row[cell_idx] if cell_idx < len(row) else None
            if cell_value is None or str(cell_value).strip() == '':
                continue

            for token in str(cell_value).strip().split():
                token = token.lower().strip()
                if token in VALID_GRADES:
                    grades_list.append({
                        'student_name': student_name,
                        'class': class_name,
                        'subject': subject,
                        'grade': token,
                        'date': grade_date_str,
                    })

    return {
        'grades': grades_list,
        'skipped': skipped,
        'period_start': period_start,
        'period_end': period_end,
        'count': len(grades_list),
    }


def generate_template_excel(class_name: str, students: list, subjects: list) -> bytes:
    """
    Генерирует Excel-шаблон для заполнения оценок.

    Строка 1: метка "Период:" + две ячейки с датами начала и конца (предзаполнены).
    Строка 2: заголовки — Предмет | Ученик1 | Ученик2 | ...
    Строки 3+: предметы, ячейки = оценки через пробел.

    Returns:
        bytes: содержимое .xlsx файла
    """
    from datetime import timedelta
    wb = Workbook()
    ws = wb.active
    ws.title = class_name

    # Стили
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    period_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    period_label_font = Font(color="FFFFFF", bold=True, size=10)
    period_date_font = Font(color="FFFFFF", bold=True, size=11)
    subj_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    subj_font = Font(color="1E40AF", bold=True, size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style='thin', color='93C5FD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_right = Side(style='medium', color='1E40AF')
    subj_border = Border(left=thin, right=thick_right, top=thin, bottom=thin)

    # ── Строка 1: период ────────────────────────────────────────────────────
    # Вычисляем текущий двухнедельный период
    now = datetime.now()
    year = now.year if now.month >= 9 else now.year - 1
    period_start = datetime(year, 9, 1)
    while period_start + timedelta(weeks=2) <= now:
        period_start += timedelta(weeks=2)
    period_end = period_start + timedelta(weeks=2) - timedelta(days=1)

    c = ws.cell(1, 1, "Период:")
    c.font = period_label_font
    c.fill = period_fill
    c.alignment = left
    c.border = border

    c = ws.cell(1, 2, period_start.strftime('%d.%m.%Y'))
    c.font = period_date_font
    c.fill = period_fill
    c.alignment = center
    c.border = border

    c = ws.cell(1, 3, period_end.strftime('%d.%m.%Y'))
    c.font = period_date_font
    c.fill = period_fill
    c.alignment = center
    c.border = border

    ws.row_dimensions[1].height = 24

    # ── Строка 2: заголовки ──────────────────────────────────────────────────
    c = ws.cell(2, 1, "Предмет")
    c.font = header_font
    c.fill = header_fill
    c.alignment = center
    c.border = border
    ws.column_dimensions['A'].width = 28

    for i, student in enumerate(students):
        col = i + 2
        c = ws.cell(2, col, student)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border
        col_letter = ws.cell(2, col).column_letter
        ws.column_dimensions[col_letter].width = 16

    ws.row_dimensions[2].height = 36

    # ── Строки 3+: предметы ──────────────────────────────────────────────────
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for row_i, subject in enumerate(subjects):
        row = row_i + 3
        c = ws.cell(row, 1, subject)
        c.font = subj_font
        c.fill = subj_fill
        c.alignment = left
        c.border = subj_border
        fill = alt_fill if row_i % 2 != 0 else None
        for i in range(len(students)):
            col = i + 2
            c = ws.cell(row, col)
            c.alignment = center
            c.border = border
            if fill:
                c.fill = fill

    # Инструкция под таблицей
    note_row = len(subjects) + 4
    c = ws.cell(note_row, 1, "Оценки: 1 2 3 4 5 н б  ·  несколько через пробел: 5 4 3")
    c.font = Font(color="64748B", italic=True, size=9)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _parse_date(value) -> Optional[datetime]:
    """Попытаться распарсить дату. Принимает datetime-объект или строку."""
    if isinstance(value, datetime):
        return value
    if not isinstance(value, str):
        value = str(value).strip()
    else:
        value = value.strip()
    for fmt in ('%d.%m.%Y', '%d.%m.%y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _match_name(raw: str, valid_lower: dict) -> Optional[str]:
    """
    Точное (case-insensitive) или нечёткое совпадение имени.
    Порог похожести: 0.82 — позволяет исправлять опечатки в 1-2 символа.
    """
    key = raw.lower().strip()
    # Сначала пробуем точное совпадение
    if key in valid_lower:
        return valid_lower[key]
    # Fuzzy-поиск через difflib
    matches = difflib.get_close_matches(key, valid_lower.keys(), n=1, cutoff=0.82)
    if matches:
        matched_key = matches[0]
        logger.debug("Fuzzy matched %r → %r", raw, valid_lower[matched_key])
        return valid_lower[matched_key]
    return None
